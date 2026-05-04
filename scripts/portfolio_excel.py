"""
포트폴리오 Excel 리포트 생성기
Usage: python scripts/portfolio_excel.py <holdings_json_path>

holdings_json 형식:
{
  "total_value": 137906438,
  "holdings": [
    {"name": "삼성전자", "value": 29350900},
    ...
  ]
}
"""
from __future__ import annotations
import json
import sys
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent.parent
TARGET_PATH = BASE_DIR / "data" / "target_portfolio.json"
OUTPUT_DIR = BASE_DIR / "output"

# ── 색상 상수 ──────────────────────────────────────────────
C_HEADER_BG  = "4472C4"   # 헤더 배경 (파란색)
C_HEADER_FG  = "FFFFFF"   # 헤더 글자 (흰색)
C_BUY_FG     = "2E75B6"   # 매수 필요 (파란 계열)
C_SELL_FG    = "C00000"   # 매도 필요 (빨강)
C_KR_BG      = "EBF3FB"   # 한국 종목 배경
C_US_BG      = "FFF2CC"   # 미국 종목 배경
C_TOTAL_BG   = "D9D9D9"   # 합계 행 배경
C_CASH_BG    = "E2EFDA"   # 현금 배경

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _match_name(extracted: str, target: str) -> bool:
    """종목명 매칭 (부분 일치 허용)"""
    e = extracted.replace(" ", "").upper()
    t = target.replace(" ", "").upper()
    return e == t or t in e or e in t


def load_target() -> dict:
    with open(TARGET_PATH, encoding="utf-8") as f:
        return json.load(f)


def build_rows(holdings: list[dict], total_value: float, target_data: dict) -> list[dict]:
    target_stocks = target_data["stocks"]
    total_inv = target_data["total_investment"]

    # holdings → name:value 맵
    held_map: dict[str, float] = {}
    for h in holdings:
        held_map[h["name"]] = h.get("value", 0)

    rows = []
    for ts in target_stocks:
        tname = ts["name"]
        target_pct = ts["target_pct"]
        country = ts["country"]

        # 보유 금액 매칭
        current_value = 0.0
        for hname, hval in held_map.items():
            if _match_name(hname, tname):
                current_value = hval
                break

        current_pct = (current_value / total_value * 100) if total_value > 0 else 0
        target_amount = total_inv * target_pct / 100
        adj_amount = target_amount - current_value
        adj_pct = target_pct - current_pct

        rows.append({
            "country": country,
            "name": tname,
            "adj_amount": adj_amount,
            "adj_pct": adj_pct,
            "target_pct": target_pct,
            "current_pct": current_pct,
            "current_value": current_value,
        })

    # 목표비중 미등록 보유 종목 추가
    registered_names = [ts["name"] for ts in target_stocks]
    for h in holdings:
        matched = any(_match_name(h["name"], rn) for rn in registered_names)
        if not matched:
            cv = h.get("value", 0)
            cp = (cv / total_value * 100) if total_value > 0 else 0
            rows.append({
                "country": "?",
                "name": h["name"] + " ⚠목표미설정",
                "adj_amount": -cv,
                "adj_pct": -cp,
                "target_pct": 0.0,
                "current_pct": cp,
                "current_value": cv,
            })

    return rows


def _cell_style(ws, row, col, value, *, bold=False, bg=None, fg=None,
                align="center", num_fmt=None, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="맑은 고딕", bold=bold, color=fg or "000000", size=10)
    c.alignment = Alignment(horizontal=align, vertical="center")
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if num_fmt:
        c.number_format = num_fmt
    if border:
        c.border = BORDER
    return c


def generate_excel(holdings: list[dict], total_value: float) -> Path:
    target_data = load_target()
    rows = build_rows(holdings, total_value, target_data)

    wb = Workbook()
    ws = wb.active
    ws.title = "포트폴리오 현황"

    # ── 투자금액 요약 (우상단) ──────────────────────────────
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 14
    ws.merge_cells("I1:J1")
    ws.cell(row=1, column=9, value="투자금액 (기준)").font = Font(name="맑은 고딕", bold=True, size=10)
    ws.cell(row=1, column=9).alignment = Alignment(horizontal="center")
    ws.merge_cells("I2:J2")
    c = ws.cell(row=1+1, column=9, value=target_data["total_investment"])
    c.font = Font(name="맑은 고딕", bold=True, size=11, color="1F4E79")
    c.alignment = Alignment(horizontal="center")
    c.number_format = '#,##0"원"'

    ws.merge_cells("I3:J3")
    ws.cell(row=3, column=9, value="현재 평가금액").font = Font(name="맑은 고딕", size=9, color="595959")
    ws.cell(row=3, column=9).alignment = Alignment(horizontal="center")
    ws.merge_cells("I4:J4")
    c = ws.cell(row=4, column=9, value=total_value)
    c.font = Font(name="맑은 고딕", bold=True, size=11, color="1F4E79")
    c.alignment = Alignment(horizontal="center")
    c.number_format = '#,##0"원"'

    # ── 컬럼 너비 ──────────────────────────────────────────
    widths = [6, 16, 14, 12, 9, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── 헤더 (6행) ─────────────────────────────────────────
    HDR_ROW = 6
    ws.row_dimensions[HDR_ROW].height = 24
    headers = ["구분", "종목", "조정 필요 금액", "조정 필요 비중", "목표비중", "현재 보유비중", "매입금액"]
    for col, h in enumerate(headers, 1):
        _cell_style(ws, HDR_ROW, col, h,
                    bold=True, bg=C_HEADER_BG, fg=C_HEADER_FG)

    # ── 데이터 행 ──────────────────────────────────────────
    data_row = HDR_ROW + 1
    for r in rows:
        ws.row_dimensions[data_row].height = 18
        country = r["country"]
        bg = C_KR_BG if country == "한국" else (C_US_BG if country == "미국" else C_CASH_BG)

        # 조정 필요 금액 색상
        adj_a = r["adj_amount"]
        adj_fg = C_BUY_FG if adj_a > 0 else (C_SELL_FG if adj_a < 0 else "595959")
        bold_row = abs(r["adj_pct"]) >= 1.0

        _cell_style(ws, data_row, 1, country, bg=bg)
        _cell_style(ws, data_row, 2, r["name"], bg=bg, align="left", bold=bold_row)

        c = _cell_style(ws, data_row, 3,
                        None if abs(adj_a) < 1 else round(adj_a),
                        fg=adj_fg, bold=bold_row, bg=bg, align="right")
        if abs(adj_a) >= 1:
            c.number_format = '#,##0'

        c = _cell_style(ws, data_row, 4,
                        round(r["adj_pct"], 2) / 100,
                        fg=adj_fg, bold=bold_row, bg=bg)
        c.number_format = '0.00%'

        c = _cell_style(ws, data_row, 5, r["target_pct"] / 100, bg=bg)
        c.number_format = '0.0%'

        c = _cell_style(ws, data_row, 6, round(r["current_pct"], 2) / 100, bg=bg)
        c.number_format = '0.00%'

        c = _cell_style(ws, data_row, 7,
                        None if r["current_value"] < 1 else round(r["current_value"]),
                        bg=bg, align="right")
        if r["current_value"] >= 1:
            c.number_format = '#,##0'

        data_row += 1

    # ── 합계 행 ────────────────────────────────────────────
    ws.row_dimensions[data_row].height = 20
    _cell_style(ws, data_row, 1, "", bold=True, bg=C_TOTAL_BG)
    _cell_style(ws, data_row, 2, "합 계", bold=True, bg=C_TOTAL_BG, align="center")
    _cell_style(ws, data_row, 3, "", bg=C_TOTAL_BG)
    _cell_style(ws, data_row, 4, "", bg=C_TOTAL_BG)
    c = _cell_style(ws, data_row, 5, 1.0, bold=True, bg=C_TOTAL_BG)
    c.number_format = '0%'
    total_current_pct = sum(r["current_pct"] for r in rows) / 100
    c = _cell_style(ws, data_row, 6, round(total_current_pct, 4), bold=True, bg=C_TOTAL_BG)
    c.number_format = '0.00%'
    c = _cell_style(ws, data_row, 7, round(total_value), bold=True, bg=C_TOTAL_BG, align="right")
    c.number_format = '#,##0'

    # ── 저장 ───────────────────────────────────────────────
    OUTPUT_DIR.mkdir(exist_ok=True)
    today = datetime.now().strftime("%Y%m%d_%H%M")
    out_path = OUTPUT_DIR / f"portfolio_status_{today}.xlsx"
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python portfolio_excel.py <holdings_json_path>")
        sys.exit(1)

    with open(sys.argv[1], encoding="utf-8") as f:
        data = json.load(f)

    out = generate_excel(data["holdings"], data["total_value"])
    print(f"[완료] {out}")
