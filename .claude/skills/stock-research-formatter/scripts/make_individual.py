"""
개별 모드: 종목 한 개당 한 파일.
이미지처럼 한 페이지 요약 표로 출력.
증권사 평가는 평가 1건당 한 행씩 펼쳐서 보기 편하게.

사용법:
    python make_individual.py <md_path> [output_path]
"""
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.insert(0, str(Path(__file__).parent))
from parse_md import parse_md


COLOR_LABEL_BG = "D9E8D4"
COLOR_HEADER_BG = "7BAE7F"
COLOR_HEADER_TEXT = "FFFFFF"
COLOR_BORDER = "B0B0B0"
COLOR_UP = "CC0000"
COLOR_DOWN = "0066CC"
COLOR_PLACEHOLDER = "999999"

FONT_NAME = "맑은 고딕"

DISPLAY_ORDER = [
    ("기업 개요", "기업 개요"),
    ("주간 성과", "주간 성과"),
    ("주간 주가 변동 주된 사유", "주간 주가 변동 주된 사유"),
    ("주가의 위치 (단기)", "주가의 위치 (단기)"),
    ("주가의 위치 (중장기)", "주가의 위치 (중장기)"),
    ("기업 경쟁력", "기업 경쟁력"),
    ("포워드 밸류에이션 추정", "포워드 밸류에이션 추정"),
    ("최근 일주일간 기업펀더멘털 변화", "최근 일주일간\n기업펀더멘털 변화"),
    ("향후 전망", "향후 전망"),
]


def _strip_md(text: str) -> str:
    """마크다운 강조 기호(**,*,__,_) 제거."""
    import re
    return re.sub(r'\*{1,2}|_{1,2}', '', text)


def _thin_border() -> Border:
    side = Side(style="thin", color=COLOR_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _detect_perf_color(value: str) -> str:
    v = value.strip().replace(" ", "")
    if v.startswith(("-", "−")):
        return COLOR_DOWN
    if v.startswith("+"):
        return COLOR_UP
    try:
        num = float(v.rstrip("%"))
        if num > 0:
            return COLOR_UP
        if num < 0:
            return COLOR_DOWN
    except ValueError:
        pass
    return "000000"


def make_individual_xlsx(source, output_path=None):
    if isinstance(source, dict):
        parsed = source
        if output_path is None:
            raise ValueError("output_path is required when source is a dict")
        output_path = Path(output_path)
    else:
        parsed = parse_md(source)
        if output_path is None:
            output_path = Path(source).with_suffix(".xlsx")
        else:
            output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = parsed["ticker"] or "리서치"

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 75

    company = parsed["company"] or "-"
    ticker = parsed["ticker"] or "-"
    date = parsed["date"] or "-"

    # 헤더
    ws.merge_cells("A1:B1")
    ws["A1"] = f"💡 학습 기업  |  {company} ({ticker})    ·    {date}"
    ws["A1"].font = Font(name=FONT_NAME, size=14, bold=True, color=COLOR_HEADER_TEXT)
    ws["A1"].fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    border = _thin_border()
    label_fill = PatternFill("solid", start_color=COLOR_LABEL_BG)
    label_font = Font(name=FONT_NAME, size=11, bold=True)
    body_font = Font(name=FONT_NAME, size=11)
    placeholder_font = Font(name=FONT_NAME, size=11, italic=True, color=COLOR_PLACEHOLDER)

    # 데이터 행
    row = 2
    for field_key, display_label in DISPLAY_ORDER:
        value = parsed["fields"].get(field_key, "")

        cell_a = ws.cell(row=row, column=1, value=display_label)
        cell_a.font = label_font
        cell_a.fill = label_fill
        cell_a.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_a.border = border

        cell_b = ws.cell(row=row, column=2)
        if value:
            cell_b.value = f"• {_strip_md(value)}"
            if field_key == "주간 성과":
                cell_b.font = Font(name=FONT_NAME, size=12, bold=True, color=_detect_perf_color(value))
            else:
                cell_b.font = body_font
        else:
            cell_b.value = "✏️ 직접 입력"
            cell_b.font = placeholder_font
        cell_b.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        cell_b.border = border

        ws.row_dimensions[row].height = max(28, min(60, len(value) // 2 + 22))
        row += 1

    # 증권사 평가 - 1건당 1행으로 펼침
    ratings = parsed["analyst_ratings"]
    if ratings:
        start_row = row
        end_row = row + len(ratings) - 1

        # 라벨 세로 병합
        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)

        cell_a = ws.cell(row=start_row, column=1, value="주간 증권사 평가")
        cell_a.font = label_font
        cell_a.fill = label_fill
        cell_a.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_a.border = border

        # 병합 영역 내 모든 셀에 배경/테두리 (병합해도 개별 적용 필요)
        for r in range(start_row, end_row + 1):
            c = ws.cell(row=r, column=1)
            c.fill = label_fill
            c.border = border

        # 각 평가 한 행씩
        for i, rating in enumerate(ratings):
            r = start_row + i
            cell_b = ws.cell(row=r, column=2, value=f"• {_strip_md(rating)}")
            cell_b.font = body_font
            cell_b.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
            cell_b.border = border
            ws.row_dimensions[r].height = 26

        row = end_row + 1
    else:
        cell_a = ws.cell(row=row, column=1, value="주간 증권사 평가")
        cell_a.font = label_font
        cell_a.fill = label_fill
        cell_a.alignment = Alignment(horizontal="center", vertical="center")
        cell_a.border = border

        cell_b = ws.cell(row=row, column=2, value="✏️ 직접 입력")
        cell_b.font = placeholder_font
        cell_b.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell_b.border = border
        ws.row_dimensions[row].height = 28
        row += 1

    # 주요 뉴스 - 1건당 1행
    news = parsed.get("news_items", [])
    if news:
        start_row = row
        end_row = row + len(news) - 1

        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)

        cell_a = ws.cell(row=start_row, column=1, value="주요 뉴스")
        cell_a.font = label_font
        cell_a.fill = label_fill
        cell_a.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_a.border = border

        for r in range(start_row, end_row + 1):
            c = ws.cell(row=r, column=1)
            c.fill = label_fill
            c.border = border

        for i, item in enumerate(news):
            r = start_row + i
            if isinstance(item, dict):
                title = item.get("title", "")
                summary = item.get("summary", "")
            else:
                title, summary = str(item), ""
            content = f"◆ {title}\n   {summary}" if summary else f"◆ {title}"
            cell_b = ws.cell(row=r, column=2, value=content)
            cell_b.font = body_font
            cell_b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
            cell_b.border = border
            ws.row_dimensions[r].height = max(40, len(summary) // 3 + 30)

        row = end_row + 1

    # 누락 경고
    if parsed["missing_fields"]:
        warn_row = row + 1
        ws.merge_cells(start_row=warn_row, start_column=1, end_row=warn_row, end_column=2)
        warn = ws.cell(row=warn_row, column=1)
        warn.value = "⚠️ 다음 항목은 사용자 입력이 필요합니다: " + ", ".join(parsed["missing_fields"])
        warn.font = Font(name=FONT_NAME, size=10, italic=True, color="CC6600")
        warn.alignment = Alignment(horizontal="left", vertical="center")

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_options.horizontalCentered = True

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python make_individual.py <md_path> [output_path]")
        sys.exit(1)

    md = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) >= 3 else None
    saved = make_individual_xlsx(md, out)
    print(f"Saved: {saved}")
