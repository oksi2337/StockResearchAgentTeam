"""
누적 모드: 여러 종목을 한 파일에 누적.
- Summary 시트: 한 종목 = 한 행 (요약/비교용)
- Ratings 시트: 평가 1건 = 한 행 (펼친 형태)

같은 (티커, 날짜) 조합이 들어오면 갱신, 새 조합이면 추가.

사용법:
    python make_portfolio.py <md_path> [portfolio_xlsx_path]
    python make_portfolio.py --bulk <research_dir> [portfolio_xlsx_path]
"""
import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from parse_md import parse_md


COLOR_HEADER_BG = "7BAE7F"
COLOR_HEADER_TEXT = "FFFFFF"
COLOR_BORDER = "B0B0B0"
COLOR_UP = "CC0000"
COLOR_DOWN = "0066CC"
COLOR_ZEBRA = "F5F8F2"

FONT_NAME = "맑은 고딕"

SUMMARY_SHEET = "Summary"
RATINGS_SHEET = "Ratings"

# Summary 시트 컬럼
SUMMARY_COLUMNS = [
    ("티커", 10),
    ("기업명", 16),
    ("작성일", 12),
    ("주간 성과", 12),
    ("주간 변동 사유", 28),
    ("단기 위치", 18),
    ("중장기 위치", 18),
    ("기업 경쟁력", 35),
    ("포워드 밸류에이션", 35),
    ("펀더멘털 변화", 35),
    ("향후 전망", 30),
    ("증권사 평가 건수", 12),
]

# Ratings 시트 컬럼
RATINGS_COLUMNS = [
    ("티커", 10),
    ("기업명", 16),
    ("작성일", 12),
    ("증권사", 18),
    ("투자의견", 16),
    ("의견 변경", 12),
    ("목표주가", 22),
    ("목표주가 변경", 18),
    ("원문", 60),
]

FIELD_MAP = {
    "주간 성과": "주간 성과",
    "주간 변동 사유": "주간 주가 변동 주된 사유",
    "단기 위치": "주가의 위치 (단기)",
    "중장기 위치": "주가의 위치 (중장기)",
    "기업 경쟁력": "기업 경쟁력",
    "포워드 밸류에이션": "포워드 밸류에이션 추정",
    "펀더멘털 변화": "최근 일주일간 기업펀더멘털 변화",
    "향후 전망": "향후 전망",
}


def _strip_md(text: str) -> str:
    """마크다운 강조 기호(**,*,__,_) 제거."""
    import re
    return re.sub(r'\*{1,2}|_{1,2}', '', text) if isinstance(text, str) else text


def _thin_border():
    side = Side(style="thin", color=COLOR_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _detect_perf_color(value):
    if not isinstance(value, str):
        return "000000"
    v = value.strip().replace(" ", "")
    if v.startswith(("-", "−")):
        return COLOR_DOWN
    if v.startswith("+"):
        return COLOR_UP
    try:
        num = float(v.rstrip("%"))
        if num > 0:
            return COLOR_UP
        if num < 0:
            return COLOR_DOWN
    except ValueError:
        pass
    return "000000"


def _parse_rating_line(rating_text: str) -> dict:
    """
    증권사 평가 한 줄을 파싱해서 컬럼별로 분리.
    예시:
        [TSM] (BofA) 투자의견, Buy (유지) + 목표주가, NTD 2,530.00 ← NTD 2,360.00 (상향)
        [TSM] (DA Davidson) 투자의견, Buy (유지) + 목표주가, $450.00 (유지)
        [NVDA] (JPMorgan) 투자의견, Overweight (상향) + 목표주가, $1,180 ← $1,050 (상향)
    """
    result = {
        "broker": "",
        "rating": "",
        "rating_change": "",
        "target_price": "",
        "target_change": "",
        "raw": rating_text,
    }

    # 증권사: 첫 괄호 안
    m = re.search(r"\(([^)]+)\)", rating_text)
    if m:
        result["broker"] = m.group(1).strip()

    # 투자의견 부분: "투자의견, XXX (변경)"
    m = re.search(r"투자의견\s*[,，]\s*([A-Za-z가-힣]+)\s*\(([^)]+)\)", rating_text)
    if m:
        result["rating"] = m.group(1).strip()
        result["rating_change"] = m.group(2).strip()

    # 목표주가 부분: "목표주가, ... (변경)"
    m = re.search(r"목표주가\s*[,，]\s*(.+?)(?:\s*\(([^)]+)\))?\s*$", rating_text)
    if m:
        price_part = m.group(1).strip()
        result["target_change"] = (m.group(2) or "").strip()

        # 화살표(←) 있으면 신/구 분리, 없으면 그대로
        if "←" in price_part:
            new_p, old_p = price_part.split("←", 1)
            result["target_price"] = f"{new_p.strip()} ← {old_p.strip()}"
        else:
            result["target_price"] = price_part

    return result


def _summary_row(parsed: dict) -> list:
    f = parsed["fields"]
    return [
        parsed["ticker"],
        parsed["company"],
        parsed["date"],
        _strip_md(f.get(FIELD_MAP["주간 성과"], "")),
        _strip_md(f.get(FIELD_MAP["주간 변동 사유"], "")),
        _strip_md(f.get(FIELD_MAP["단기 위치"], "")),
        _strip_md(f.get(FIELD_MAP["중장기 위치"], "")),
        _strip_md(f.get(FIELD_MAP["기업 경쟁력"], "")),
        _strip_md(f.get(FIELD_MAP["포워드 밸류에이션"], "")),
        _strip_md(f.get(FIELD_MAP["펀더멘털 변화"], "")),
        _strip_md(f.get(FIELD_MAP["향후 전망"], "")),
        len(parsed.get("analyst_ratings", [])),
    ]


def _rating_rows(parsed: dict) -> list[list]:
    """파싱 결과의 analyst_ratings를 Ratings 시트 행 리스트로 변환."""
    rows = []
    for line in parsed.get("analyst_ratings", []):
        info = _parse_rating_line(_strip_md(line))
        rows.append([
            parsed["ticker"],
            parsed["company"],
            parsed["date"],
            info["broker"],
            info["rating"],
            info["rating_change"],
            info["target_price"],
            info["target_change"],
            _strip_md(info["raw"]),
        ])
    return rows


def _build_header(ws, columns):
    """헤더 1행 작성 + 컬럼 너비 설정 + freeze pane."""
    border = _thin_border()
    header_fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
    header_font = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_HEADER_TEXT)

    for idx, (name, width) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
        c = ws.cell(row=1, column=idx, value=name)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"


def _init_workbook(path: Path) -> Workbook:
    """파일이 없으면 새 워크북(2개 시트), 있으면 로드. 시트 보장."""
    if path.exists():
        wb = load_workbook(path)
        # 누적 시트 보장
        if SUMMARY_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(SUMMARY_SHEET, 0)
            _build_header(ws, SUMMARY_COLUMNS)
        if RATINGS_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(RATINGS_SHEET)
            _build_header(ws, RATINGS_COLUMNS)
        return wb

    wb = Workbook()
    # 기본 시트를 Summary로 이름 변경
    ws_summary = wb.active
    ws_summary.title = SUMMARY_SHEET
    _build_header(ws_summary, SUMMARY_COLUMNS)

    ws_ratings = wb.create_sheet(RATINGS_SHEET)
    _build_header(ws_ratings, RATINGS_COLUMNS)

    return wb


def _write_summary_row(ws, row_idx: int, row_data: list):
    border = _thin_border()
    body_font = Font(name=FONT_NAME, size=10)
    zebra_fill = PatternFill("solid", start_color=COLOR_ZEBRA) if row_idx % 2 == 0 else None

    for col_idx, value in enumerate(row_data, start=1):
        c = ws.cell(row=row_idx, column=col_idx, value=value if value != "" else None)

        if col_idx == 4 and isinstance(value, str) and value:
            c.font = Font(name=FONT_NAME, size=10, bold=True, color=_detect_perf_color(value))
        else:
            c.font = body_font

        c.alignment = Alignment(
            horizontal="center" if col_idx in (1, 2, 3, 4, 12) else "left",
            vertical="center",
            wrap_text=True,
        )
        c.border = border
        if zebra_fill:
            c.fill = zebra_fill

    ws.row_dimensions[row_idx].height = 36


def _write_rating_row(ws, row_idx: int, row_data: list):
    border = _thin_border()
    body_font = Font(name=FONT_NAME, size=10)
    zebra_fill = PatternFill("solid", start_color=COLOR_ZEBRA) if row_idx % 2 == 0 else None

    for col_idx, value in enumerate(row_data, start=1):
        c = ws.cell(row=row_idx, column=col_idx, value=value if value != "" else None)
        c.font = body_font
        # 변경 컬럼(상향/하향) 색상 강조
        if col_idx in (6, 8) and isinstance(value, str):
            if "상향" in value:
                c.font = Font(name=FONT_NAME, size=10, bold=True, color=COLOR_UP)
            elif "하향" in value:
                c.font = Font(name=FONT_NAME, size=10, bold=True, color=COLOR_DOWN)

        c.alignment = Alignment(
            horizontal="center" if col_idx in (1, 2, 3, 4, 5, 6, 8) else "left",
            vertical="center",
            wrap_text=True,
        )
        c.border = border
        if zebra_fill:
            c.fill = zebra_fill

    ws.row_dimensions[row_idx].height = 28


def _find_summary_row(ws, ticker: str, date: str):
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == ticker and ws.cell(row=r, column=3).value == date:
            return r
    return None


def _delete_rating_rows_for(ws, ticker: str, date: str):
    """기존 (티커, 날짜) 평가 행을 모두 삭제 (갱신 위해)."""
    rows_to_delete = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == ticker and ws.cell(row=r, column=3).value == date:
            rows_to_delete.append(r)

    # 뒤에서부터 삭제 (인덱스 꼬임 방지)
    for r in reversed(rows_to_delete):
        ws.delete_rows(r)


def add_to_portfolio(source, portfolio_path) -> tuple:
    """단일 md 또는 dict를 포트폴리오에 추가/갱신. (저장경로, 'added'|'updated', 평가건수) 반환."""
    portfolio_path = Path(portfolio_path)
    portfolio_path.parent.mkdir(parents=True, exist_ok=True)

    parsed = source if isinstance(source, dict) else parse_md(Path(source))
    ticker = parsed["ticker"]
    date = parsed["date"]

    wb = _init_workbook(portfolio_path)

    # ── Summary 시트 ──
    ws_s = wb[SUMMARY_SHEET]
    summary = _summary_row(parsed)
    existing = _find_summary_row(ws_s, ticker, date)
    if existing:
        _write_summary_row(ws_s, existing, summary)
        action = "updated"
    else:
        _write_summary_row(ws_s, ws_s.max_row + 1, summary)
        action = "added"

    # ── Ratings 시트 ──
    ws_r = wb[RATINGS_SHEET]
    # 기존 평가 행 제거 (갱신 시)
    _delete_rating_rows_for(ws_r, ticker, date)

    # 새 평가 행 추가
    rating_rows = _rating_rows(parsed)
    for rd in rating_rows:
        _write_rating_row(ws_r, ws_r.max_row + 1, rd)

    wb.save(portfolio_path)
    return portfolio_path, action, len(rating_rows)


def add_bulk(research_dir, portfolio_path) -> tuple:
    research_dir = Path(research_dir)
    md_files = sorted(research_dir.glob("*.md"))

    log = []
    for md in md_files:
        try:
            _, action, cnt = add_to_portfolio(md, portfolio_path)
            log.append((md.name, action, cnt))
        except Exception as e:
            log.append((md.name, f"error: {e}", 0))

    return Path(portfolio_path), log


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python make_portfolio.py <md_path> [portfolio_xlsx]")
        print("  python make_portfolio.py --bulk <research_dir> [portfolio_xlsx]")
        sys.exit(1)

    if sys.argv[1] == "--bulk":
        research = sys.argv[2]
        out = sys.argv[3] if len(sys.argv) >= 4 else "portfolio_master.xlsx"
        saved, log = add_bulk(research, out)
        print(f"Saved: {saved}")
        for entry in log:
            if len(entry) == 3:
                name, action, cnt = entry
                print(f"  {action}: {name} (평가 {cnt}건)")
            else:
                print(f"  {entry}")
    else:
        md = sys.argv[1]
        out = sys.argv[2] if len(sys.argv) >= 3 else "portfolio_master.xlsx"
        saved, action, cnt = add_to_portfolio(md, out)
        print(f"Saved: {saved} ({action}, 평가 {cnt}건)")
